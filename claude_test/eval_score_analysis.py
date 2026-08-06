# Intent: one-off analysis of outputs/evaluation/data.xlsx for gh #140. The
# workbook only carries mean / SD / CV per condition; this turns the 135 scored
# trials into per-condition component rates with Wilson intervals, a failure-mode
# breakdown of the Korean 비고 notes, and the report figures. Expected to live
# until the evaluation is superseded by a re-run -- regenerate its outputs rather
# than editing them by hand. The workbook itself is never written to (#132).

import csv
import math
import re
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

repo_root = Path(__file__).resolve().parent.parent
workbook_path = repo_root / "outputs" / "evaluation" / "data.xlsx"
video_report_path = repo_root / "outputs" / "evaluation" / "analysis" / "README.md"
output_dir = repo_root / "outputs" / "evaluation" / "analysis" / "scores"

# The Summary sheet labels these 작업 1/2/3, and the recordings and Hub
# checkpoints use the matching task1/2/3 prefixes.
sheet_tasks = {
    "Pick and Place": ("task1", "Pick and Place"),
    "Pouring": ("task2", "Pouring"),
    "밸브 조작": ("task3", "Valve"),
}
task_order = ["task1", "task2", "task3"]
task_titles = {"task1": "Pick and Place", "task2": "Pouring", "task3": "Valve"}
model_order = ["ACT", "Pi0", "Pi0.5"]
dataset_order = [50, 100, 200]
trials_per_condition = 5
max_score = 10.0

# Rubric printed on every raw sheet: 2 points for finishing inside the time
# limit, 4 for task success, 4 for two consecutive successes.
weight_in_time = 2
weight_success = 4
weight_consecutive = 4

# Summary sheet block layout: one row per model, three columns per dataset size.
summary_first_rows = {"task1": 8, "task2": 15, "task3": 22}
summary_mean_columns = {50: 3, 100: 6, 200: 9}

# dataviz reference palette, light surface. Slots 1-3 are the set the reference
# certifies for all-pairs CVD separation, so three models fit without re-stepping.
color_by_model = {"ACT": "#2a78d6", "Pi0": "#eb6834", "Pi0.5": "#1baf7a"}
surface = "#fcfcfb"
ink_primary = "#0b0b0b"
ink_secondary = "#52514e"
ink_muted = "#898781"
gridline = "#e1e0d9"
baseline = "#c3c2b7"
sequential_blue = [
    "#cde2fb",
    "#b7d3f6",
    "#9ec5f4",
    "#86b6ef",
    "#6da7ec",
    "#5598e7",
    "#3987e5",
    "#2a78d6",
    "#256abf",
    "#1c5cab",
    "#184f95",
    "#104281",
    "#0d366b",
]

# Each rule tags a note fragment. A fragment may match several rules -- a single
# comment often describes more than one thing -- so these are labels, not bins.
# `operator_event` is orthogonal: it marks rig or operator interference rather
# than policy behaviour, and must not be counted as a policy failure.
note_rules = [
    ("operator_event", r"무효|재부팅|수리함|재시[가작]|카메라선 손상|녹화함|모터 부하"),
    ("motion_instability", r"모션이 튀|모션이 튐|진동만"),
    ("gripper_close_failure", r"그리퍼를 못 닫"),
    ("grasp_failure", r"못\s*집|못\s*잡|못 듬|못 뺌|들진 못함|집으려[^,]*실패"),
    ("release_failure", r"못 놓았"),
    ("drop", r"떨어트|놓침"),
    ("collision_or_push", r"찍음|부딪|부딧|주변을 침|누름|밈|세게"),
    ("premature_home_return", r"원점으로\s*돌아|중간에 복귀"),
    ("no_termination", r"계속 부"),
    ("incomplete_pour", r"못 부음|못 부"),
    ("return_failure", r"반납 못함|복귀 못함"),
    ("placement_failure", r"못\s*넘|못 넣|못 이동"),
    ("hesitation", r"머뭇거|멈칫|멈춰|버벅|걸림"),
    ("misperception", r"위치를 잘 못 인식|이미 밸브가 돌아가"),
    ("partial_progress", r"집기만 함|집지만|집었지만"),
    ("positive_remark", r"성공함|잘함|정상 수행|아슬아슬|더 잘|테크닉"),
]
policy_failure_modes = [
    "grasp_failure",
    "collision_or_push",
    "drop",
    "placement_failure",
    "motion_instability",
    "return_failure",
    "hesitation",
    "premature_home_return",
    "no_termination",
    "incomplete_pour",
    "misperception",
    "gripper_close_failure",
    "release_failure",
    "partial_progress",
]
failure_mode_labels = {
    "grasp_failure": "Grasp failure",
    "collision_or_push": "Collision / push",
    "drop": "Dropped object",
    "placement_failure": "Placement failure",
    "motion_instability": "Motion instability",
    "return_failure": "Return failure",
    "hesitation": "Hesitation / stall",
    "premature_home_return": "Premature home return",
    "no_termination": "No termination",
    "incomplete_pour": "Incomplete pour",
    "misperception": "Misperception",
    "gripper_close_failure": "Gripper close failure",
    "release_failure": "Release failure",
    "partial_progress": "Partial progress only",
}


def load_trials():
    """Read the three raw sheets into one flat list of trial records."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    trials = []
    for sheet_name, (task, task_label) in sheet_tasks.items():
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=6, max_row=50, min_col=2, max_col=9):
            model = row[0].value
            if model is None:
                continue
            in_time = bool(row[3].value)
            success = bool(row[4].value)
            consecutive = bool(row[5].value)
            trials.append(
                {
                    "task": task,
                    "task_label": task_label,
                    "sheet": sheet_name,
                    "sheet_row": row[0].row,
                    "model": str(model),
                    "dataset": int(row[1].value),
                    "trial": str(row[2].value),
                    "in_time": in_time,
                    "success": success,
                    "consecutive": consecutive,
                    "score_recorded": float(row[6].value),
                    "score_rubric": float(
                        weight_in_time * in_time + weight_success * success + weight_consecutive * consecutive
                    ),
                    "note": (row[7].value or "").strip(),
                }
            )
    return trials


def read_summary_means(workbook):
    """Pull the per-condition means the workbook already publishes."""
    sheet = workbook["Summary"]
    means = {}
    for task, first_row in summary_first_rows.items():
        for offset, model in enumerate(model_order):
            for dataset, column in summary_mean_columns.items():
                value = sheet.cell(row=first_row + offset, column=column).value
                means[(task, model, dataset)] = float(value)
    return means


def wilson_interval(successes, total, z=1.959963985):
    """Return the Wilson score interval for a proportion.

    Chosen over the coefficient of variation the workbook uses because it stays
    defined at 0/5 and 5/5, where the CV is not.
    """
    if total == 0:
        return (0.0, 0.0)
    proportion = successes / total
    denominator = 1 + z * z / total
    center = (proportion + z * z / (2 * total)) / denominator
    spread = z * math.sqrt(proportion * (1 - proportion) / total + z * z / (4 * total * total)) / denominator
    return (max(0.0, center - spread), min(1.0, center + spread))


def classify_note(note):
    """Tag one 비고 note with every failure mode its fragments mention."""
    tags = set()
    matched_fragments = []
    fragments = [f.strip() for f in re.split(r"[,.]", note) if f.strip()]
    for fragment in fragments:
        fragment_tags = {name for name, pattern in note_rules if re.search(pattern, fragment)}
        if fragment_tags:
            matched_fragments.append(fragment)
        tags |= fragment_tags
    unmatched = [f for f in fragments if f not in matched_fragments]
    return tags, unmatched


def build_conditions(trials):
    """Aggregate the trials into the 27 task x model x dataset conditions."""
    grouped = defaultdict(list)
    for trial in trials:
        grouped[(trial["task"], trial["model"], trial["dataset"])].append(trial)

    conditions = {}
    for task in task_order:
        for model in model_order:
            for dataset in dataset_order:
                group = grouped[(task, model, dataset)]
                rubric = [t["score_rubric"] for t in group]
                recorded = [t["score_recorded"] for t in group]
                entry = {
                    "task": task,
                    "task_label": task_titles[task],
                    "model": model,
                    "dataset": dataset,
                    "n": len(group),
                    "mean_rubric": statistics.mean(rubric),
                    "sd_rubric": statistics.stdev(rubric),
                    "mean_recorded": statistics.mean(recorded),
                }
                for field in ("success", "in_time", "consecutive"):
                    count = sum(t[field] for t in group)
                    low, high = wilson_interval(count, len(group))
                    entry[f"{field}_count"] = count
                    entry[f"{field}_rate"] = count / len(group)
                    entry[f"{field}_ci_low"] = low
                    entry[f"{field}_ci_high"] = high
                conditions[(task, model, dataset)] = entry
    return conditions


def check_rubric(trials):
    """List the trials whose stored total disagrees with its own booleans."""
    return [t for t in trials if t["score_recorded"] != t["score_rubric"]]


def parse_video_durations():
    """Read the measured per-attempt durations from the video analysis report.

    Covers the 18 ACT / Pi0 conditions only; Pi0.5 was never run through
    ``eval_video_report.py``.
    """
    pattern = re.compile(
        r"\|\s*`(task\d)_(act|pi0)_(\d+)`\s*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|"
        r"\s*([\d.]+)\s*±\s*[\d.]+\s*\|\s*(\d+)~(\d+)\s*\|"
    )
    measured = {}
    for line in video_report_path.read_text(encoding="utf-8").splitlines():
        match = pattern.search(line)
        if match:
            task, model_slug, dataset, mean_s, min_s, max_s = match.groups()
            model = {"act": "ACT", "pi0": "Pi0"}[model_slug]
            measured[(task, model, int(dataset))] = {
                "mean_s": float(mean_s),
                "min_s": int(min_s),
                "max_s": int(max_s),
            }
    return measured


def write_trials_csv(trials):
    path = output_dir / "trials.csv"
    fields = [
        "task",
        "task_label",
        "sheet",
        "sheet_row",
        "model",
        "dataset",
        "trial",
        "in_time",
        "success",
        "consecutive",
        "score_recorded",
        "score_rubric",
        "failure_modes",
        "note",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for trial in trials:
            row = {field: trial[field] for field in fields if field != "failure_modes"}
            row["failure_modes"] = ";".join(sorted(trial["tags"]))
            writer.writerow(row)
    return path


def write_conditions_csv(conditions):
    path = output_dir / "conditions.csv"
    fields = [
        "task",
        "task_label",
        "model",
        "dataset",
        "n",
        "mean_rubric",
        "sd_rubric",
        "mean_recorded",
        "success_count",
        "success_rate",
        "success_ci_low",
        "success_ci_high",
        "in_time_count",
        "in_time_rate",
        "in_time_ci_low",
        "in_time_ci_high",
        "consecutive_count",
        "consecutive_rate",
        "consecutive_ci_low",
        "consecutive_ci_high",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for task in task_order:
            for model in model_order:
                for dataset in dataset_order:
                    writer.writerow(conditions[(task, model, dataset)])
    return path


def style_axes(axes, *, hide_left=False):
    axes.set_facecolor(surface)
    for side in ("top", "right"):
        axes.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        axes.spines[side].set_color(baseline)
        axes.spines[side].set_linewidth(1.0)
    if hide_left:
        axes.spines["left"].set_visible(False)
    axes.tick_params(colors=ink_muted, labelsize=9, length=0)
    for label in axes.get_xticklabels() + axes.get_yticklabels():
        label.set_color(ink_secondary)


def plot_heatmap(conditions):
    """Score magnitude over the whole grid -- sequential, one hue."""
    rows = [(task, model) for task in task_order for model in model_order]
    grid = np.array(
        [[conditions[(task, model, ds)]["mean_rubric"] for ds in dataset_order] for task, model in rows]
    )
    colormap = matplotlib.colors.LinearSegmentedColormap.from_list("blue_sequential", sequential_blue)
    figure, axes = plt.subplots(figsize=(6.4, 6.6), facecolor=surface)
    axes.imshow(grid, cmap=colormap, vmin=0, vmax=max_score, aspect="auto")
    axes.set_xticks(range(len(dataset_order)))
    axes.set_xticklabels([f"{ds} episodes" for ds in dataset_order])
    axes.set_yticks(range(len(rows)))
    axes.set_yticklabels([model for _, model in rows])
    for index, value in np.ndenumerate(grid):
        # The light steps of the ramp recede toward the surface, so flip the
        # label ink once the cell is dark enough to swallow black text.
        axes.text(
            index[1],
            index[0],
            f"{value:.1f}",
            ha="center",
            va="center",
            fontsize=11,
            weight="bold",
            color=surface if value > 6.5 else ink_primary,
        )
    for boundary in (2.5, 5.5):
        axes.axhline(boundary, color=surface, linewidth=3)
    for offset, task in enumerate(task_order):
        axes.text(
            -0.92,
            offset * 3 + 1,
            task_titles[task],
            rotation=90,
            ha="center",
            va="center",
            fontsize=10,
            color=ink_primary,
            weight="bold",
        )
    style_axes(axes)
    axes.set_xticks(np.arange(-0.5, len(dataset_order), 1), minor=True)
    axes.set_yticks(np.arange(-0.5, len(rows), 1), minor=True)
    axes.grid(which="minor", color=surface, linewidth=2)
    axes.tick_params(which="minor", length=0)
    axes.set_title(
        "Mean rubric score by task, model and dataset size",
        color=ink_primary,
        fontsize=12.5,
        pad=22,
        loc="left",
    )
    axes.text(
        0,
        1.015,
        "0-10 scale, mean of 5 attempts",
        fontsize=9,
        color=ink_muted,
        ha="left",
        va="bottom",
        transform=axes.transAxes,
    )
    figure.tight_layout()
    path = output_dir / "fig1_heatmap.png"
    figure.savefig(path, dpi=180, facecolor=surface)
    plt.close(figure)
    return path


def plot_scaling(conditions):
    """Score against dataset size, one panel per task."""
    figure, axes_row = plt.subplots(1, 3, figsize=(11.5, 4.2), sharey=True, facecolor=surface)
    positions = range(len(dataset_order))
    for axes, task in zip(axes_row, task_order, strict=True):
        axes.set_ylim(-0.6, 11.2)
        axes.grid(axis="y", color=gridline, linewidth=1)
        axes.set_axisbelow(True)
        for model in model_order:
            values = [conditions[(task, model, ds)]["mean_rubric"] for ds in dataset_order]
            axes.plot(
                positions,
                values,
                color=color_by_model[model],
                linewidth=2,
                marker="o",
                markersize=8,
                markeredgecolor=surface,
                markeredgewidth=2,
                label=model,
                zorder=3,
            )
            axes.annotate(
                f"{values[-1]:.1f}",
                (len(dataset_order) - 1, values[-1]),
                textcoords="offset points",
                xytext=(9, 0),
                va="center",
                fontsize=9,
                color=ink_secondary,
            )
        axes.set_xticks(list(positions))
        axes.set_xticklabels([str(ds) for ds in dataset_order])
        axes.set_xlim(-0.35, len(dataset_order) - 0.35)
        axes.set_title(task_titles[task], color=ink_primary, fontsize=11, loc="left")
        style_axes(axes)
    axes_row[0].set_ylabel("Mean rubric score", color=ink_secondary, fontsize=10)
    axes_row[1].set_xlabel("Training episodes", color=ink_secondary, fontsize=10)
    handles, labels = axes_row[0].get_legend_handles_labels()
    figure.legend(
        handles,
        labels,
        loc="upper right",
        frameon=False,
        ncol=3,
        fontsize=10,
        labelcolor=ink_secondary,
        bbox_to_anchor=(0.99, 1.005),
    )
    figure.suptitle(
        "Data scaling is neither monotonic nor uniform across tasks",
        color=ink_primary,
        fontsize=12.5,
        x=0.008,
        ha="left",
        y=0.985,
    )
    figure.tight_layout(rect=(0, 0, 1, 0.945))
    path = output_dir / "fig2_scaling.png"
    figure.savefig(path, dpi=180, facecolor=surface)
    plt.close(figure)
    return path


def plot_success_intervals(conditions):
    """Task-success rate with Wilson 95 % intervals."""
    figure, axes_row = plt.subplots(1, 3, figsize=(11.5, 4.4), sharey=True, facecolor=surface)
    bar_width = 0.26
    for axes, task in zip(axes_row, task_order, strict=True):
        axes.set_ylim(0, 1.12)
        axes.grid(axis="y", color=gridline, linewidth=1)
        axes.set_axisbelow(True)
        for model_index, model in enumerate(model_order):
            centers = [i + (model_index - 1) * (bar_width + 0.02) for i in range(len(dataset_order))]
            rates = [conditions[(task, model, ds)]["success_rate"] for ds in dataset_order]
            lows = [
                rates[i] - conditions[(task, model, ds)]["success_ci_low"]
                for i, ds in enumerate(dataset_order)
            ]
            highs = [
                conditions[(task, model, ds)]["success_ci_high"] - rates[i]
                for i, ds in enumerate(dataset_order)
            ]
            axes.bar(
                centers,
                rates,
                width=bar_width,
                color=color_by_model[model],
                label=model,
                zorder=3,
                edgecolor=surface,
                linewidth=2,
            )
            axes.errorbar(
                centers,
                rates,
                yerr=[lows, highs],
                fmt="none",
                ecolor=ink_secondary,
                elinewidth=1.2,
                capsize=3,
                zorder=4,
            )
            for center, rate, dataset in zip(centers, rates, dataset_order, strict=True):
                count = conditions[(task, model, dataset)]["success_count"]
                axes.text(
                    center,
                    0.03,
                    f"{count}/5",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                    color=surface if rate > 0.25 else ink_secondary,
                    zorder=5,
                )
        axes.set_xticks(range(len(dataset_order)))
        axes.set_xticklabels([str(ds) for ds in dataset_order])
        axes.set_title(task_titles[task], color=ink_primary, fontsize=11, loc="left")
        style_axes(axes)
    axes_row[0].set_ylabel("Task success rate", color=ink_secondary, fontsize=10)
    axes_row[0].set_yticks([0, 0.25, 0.5, 0.75, 1.0])
    axes_row[0].set_yticklabels(["0", "25%", "50%", "75%", "100%"])
    axes_row[1].set_xlabel("Training episodes", color=ink_secondary, fontsize=10)
    handles, labels = axes_row[0].get_legend_handles_labels()
    figure.legend(
        handles,
        labels,
        loc="upper right",
        frameon=False,
        ncol=3,
        fontsize=10,
        labelcolor=ink_secondary,
        bbox_to_anchor=(0.99, 1.005),
    )
    figure.suptitle(
        "Task success rate with Wilson 95 % intervals (n = 5 per condition)",
        color=ink_primary,
        fontsize=12.5,
        x=0.008,
        ha="left",
        y=0.985,
    )
    figure.tight_layout(rect=(0, 0, 1, 0.945))
    path = output_dir / "fig3_success_ci.png"
    figure.savefig(path, dpi=180, facecolor=surface)
    plt.close(figure)
    return path


def plot_components(conditions):
    """Show where the headline score is lost: time limit, success, repeatability."""
    components = [
        ("in_time_rate", "Finished in time"),
        ("success_rate", "Task success"),
        ("consecutive_rate", "Two consecutive"),
    ]
    figure, axes_grid = plt.subplots(3, 3, figsize=(11.5, 7.4), sharey=True, sharex=True, facecolor=surface)
    positions = range(len(dataset_order))
    for row_index, (field, component_label) in enumerate(components):
        for column_index, task in enumerate(task_order):
            axes = axes_grid[row_index][column_index]
            axes.set_ylim(-0.06, 1.14)
            axes.grid(axis="y", color=gridline, linewidth=1)
            axes.set_axisbelow(True)
            for model in model_order:
                values = [conditions[(task, model, ds)][field] for ds in dataset_order]
                axes.plot(
                    positions,
                    values,
                    color=color_by_model[model],
                    linewidth=2,
                    marker="o",
                    markersize=7,
                    markeredgecolor=surface,
                    markeredgewidth=2,
                    label=model,
                    zorder=3,
                )
            if row_index == 0:
                axes.set_title(task_titles[task], color=ink_primary, fontsize=11, loc="left")
            if column_index == 0:
                axes.set_ylabel(component_label, color=ink_primary, fontsize=10)
            axes.set_xticks(list(positions))
            axes.set_xticklabels([str(ds) for ds in dataset_order])
            axes.set_xlim(-0.3, len(dataset_order) - 0.7)
            axes.set_yticks([0, 0.5, 1.0])
            axes.set_yticklabels(["0", "50%", "100%"])
            style_axes(axes)
    axes_grid[2][1].set_xlabel("Training episodes", color=ink_secondary, fontsize=10)
    handles, labels = axes_grid[0][0].get_legend_handles_labels()
    figure.legend(
        handles,
        labels,
        loc="upper right",
        frameon=False,
        ncol=3,
        fontsize=10,
        labelcolor=ink_secondary,
        bbox_to_anchor=(0.99, 1.005),
    )
    figure.suptitle(
        "Rubric components separately: the total score conflates speed with success",
        color=ink_primary,
        fontsize=12.5,
        x=0.008,
        ha="left",
        y=0.99,
    )
    figure.tight_layout(rect=(0, 0, 1, 0.955))
    path = output_dir / "fig4_components.png"
    figure.savefig(path, dpi=180, facecolor=surface)
    plt.close(figure)
    return path


def plot_failure_modes(counts_by_model, ordered_modes):
    """Failure-mode composition per model, from the 비고 notes."""
    figure, axes = plt.subplots(figsize=(9.4, 6.4), facecolor=surface)
    height = 0.26
    offsets = np.arange(len(ordered_modes))
    for model_index, model in enumerate(model_order):
        centers = offsets + (model_index - 1) * (height + 0.02)
        values = [counts_by_model[model][mode] for mode in ordered_modes]
        axes.barh(
            centers,
            values,
            height=height,
            color=color_by_model[model],
            label=model,
            zorder=3,
            edgecolor=surface,
            linewidth=2,
        )
        for center, value in zip(centers, values, strict=True):
            if value:
                axes.text(
                    value + 0.4,
                    center,
                    str(value),
                    va="center",
                    fontsize=8.5,
                    color=ink_secondary,
                )
    axes.set_yticks(offsets)
    axes.set_yticklabels([failure_mode_labels[mode] for mode in ordered_modes])
    axes.invert_yaxis()
    axes.grid(axis="x", color=gridline, linewidth=1)
    axes.set_axisbelow(True)
    axes.set_xlabel("Trials mentioning the mode (of 45 per model)", color=ink_secondary, fontsize=10)
    style_axes(axes, hide_left=True)
    axes.legend(frameon=False, fontsize=10, labelcolor=ink_secondary, loc="lower right")
    axes.set_title(
        "Failure signatures differ by model, not just in frequency",
        color=ink_primary,
        fontsize=12.5,
        pad=14,
        loc="left",
    )
    figure.tight_layout()
    path = output_dir / "fig5_failure_modes.png"
    figure.savefig(path, dpi=180, facecolor=surface)
    plt.close(figure)
    return path


def format_condition_table(conditions):
    lines = [
        "| Task | Model | 50 ep | 100 ep | 200 ep |",
        "|---|---|---|---|---|",
    ]
    for task in task_order:
        for index, model in enumerate(model_order):
            cells = []
            for dataset in dataset_order:
                entry = conditions[(task, model, dataset)]
                cells.append(f"{entry['mean_rubric']:.1f} · {entry['success_count']}/5")
            label = task_titles[task] if index == 0 else ""
            lines.append(f"| {label} | {model} | " + " | ".join(cells) + " |")
    return "\n".join(lines)


def format_component_table(conditions):
    lines = [
        "| Task | Model | Dataset | In time | Success | Two consecutive | Mean score |",
        "|---|---|---|---|---|---|---|",
    ]
    for task in task_order:
        for model in model_order:
            for dataset in dataset_order:
                entry = conditions[(task, model, dataset)]
                lines.append(
                    f"| {task_titles[task]} | {model} | {dataset} | "
                    f"{entry['in_time_count']}/5 | "
                    f"{entry['success_count']}/5 "
                    f"({entry['success_ci_low'] * 100:.0f}–{entry['success_ci_high'] * 100:.0f} %) | "
                    f"{entry['consecutive_count']}/5 | {entry['mean_rubric']:.1f} |"
                )
    return "\n".join(lines)


def format_failure_table(counts_by_model, ordered_modes):
    lines = ["| Failure mode | ACT | Pi0 | Pi0.5 |", "|---|---|---|---|"]
    for mode in ordered_modes:
        counts = [counts_by_model[model][mode] for model in model_order]
        lines.append(f"| {failure_mode_labels[mode]} | " + " | ".join(str(c) for c in counts) + " |")
    return "\n".join(lines)


def main():
    output_dir.mkdir(parents=True, exist_ok=True)
    trials = load_trials()
    expected_trials = len(task_order) * len(model_order) * len(dataset_order) * trials_per_condition
    assert len(trials) == expected_trials, f"expected {expected_trials} trials, got {len(trials)}"

    unmatched_fragments = []
    for trial in trials:
        tags, unmatched = classify_note(trial["note"])
        trial["tags"] = tags
        unmatched_fragments.extend(unmatched)

    conditions = build_conditions(trials)
    assert len(conditions) == 27, f"expected 27 conditions, got {len(conditions)}"

    print("=" * 78)
    print("1. Scoring defects -- stored total vs the row's own booleans")
    print("=" * 78)
    mismatches = check_rubric(trials)
    for trial in mismatches:
        print(
            f"  {trial['sheet']} r{trial['sheet_row']}: {trial['model']}/"
            f"{trial['dataset']}/{trial['trial']} stored {trial['score_recorded']:.0f} "
            f"vs rubric {trial['score_rubric']:.0f}  |  {trial['note']}"
        )
    print(f"  {len(mismatches)} of {len(trials)} trials disagree.")

    # The recomputed means must reproduce the workbook everywhere the stored
    # totals are self-consistent. Anything else is a bug in this script.
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    published = read_summary_means(workbook)
    affected = {(t["task"], t["model"], t["dataset"]) for t in mismatches}
    for key, entry in conditions.items():
        assert math.isclose(entry["mean_recorded"], published[key], abs_tol=1e-6), (
            f"recorded mean for {key} does not reproduce the Summary sheet"
        )
        if key not in affected:
            assert math.isclose(entry["mean_rubric"], published[key], abs_tol=1e-6), (
                f"unexpected rubric divergence at {key}"
            )
    print("  Summary sheet reproduced; rubric differs only at:")
    for key in sorted(affected):
        print(f"    {key}: {published[key]:.1f} -> {conditions[key]['mean_rubric']:.1f}")

    print()
    print("=" * 78)
    print("2. Trials annotated 무효 (invalid) but still scored")
    print("=" * 78)
    invalid = [t for t in trials if "무효" in t["note"]]
    for trial in invalid:
        print(
            f"  {trial['task']} {trial['model']}/{trial['dataset']}/{trial['trial']} "
            f"score {trial['score_rubric']:.0f}  |  {trial['note']}"
        )
    print(f"  {len(invalid)} trials.")

    print()
    print("=" * 78)
    print("3. What the time-limit flag actually tracks")
    print("=" * 78)
    per_task_divergence = Counter()
    implied = [t for t in trials if t["in_time"] and not t["success"]]
    for trial in trials:
        if trial["in_time"] != trial["success"]:
            per_task_divergence[trial["task"]] += 1
    print(f"  in_time == success on {len(trials) - sum(per_task_divergence.values())}/{len(trials)} trials.")
    for task in task_order:
        print(f"    {task_titles[task]}: {per_task_divergence[task]} divergent trials.")
    print(f"  in_time True while success False: {len(implied)} trials.")

    measured = parse_video_durations()
    print(f"  Video-measured durations available for {len(measured)}/27 conditions (ACT and Pi0 only).")
    print("  condition                in_time  success  measured min~max (s)")
    for task in task_order:
        for model in ("ACT", "Pi0"):
            for dataset in dataset_order:
                key = (task, model, dataset)
                if key not in measured:
                    continue
                entry = conditions[key]
                stats = measured[key]
                print(
                    f"    {task}_{model.lower()}_{dataset:<4}       "
                    f"{entry['in_time_count']}/5      {entry['success_count']}/5     "
                    f"{stats['min_s']}~{stats['max_s']}"
                )
    # The criteria banner on each raw sheet still reads 40 s / 60 s / 40 s while
    # the column header reads 180 s. The video report gives per-attempt durations
    # but not which attempt belongs to which 회차, so neither reading can be
    # falsified outright; report what the durations do and do not rule out.
    banner_limits = {"task1": 40, "task2": 60, "task3": 40}
    strict = [
        key
        for key, stats in measured.items()
        if conditions[key]["in_time_count"] > 0 and stats["min_s"] > banner_limits[key[0]]
    ]
    print(f"  Marked in time although every measured attempt exceeded the banner limit: {len(strict)}")
    print("  Nothing penalised although some attempt exceeded the banner limit:")
    for key, stats in sorted(measured.items()):
        if (
            conditions[key]["in_time_count"] == trials_per_condition
            and stats["max_s"] > banner_limits[key[0]]
        ):
            print(
                f"    {key[0]}_{key[1].lower()}_{key[2]}: 5/5 in time, "
                f"slowest measured attempt {stats['max_s']} s vs {banner_limits[key[0]]} s banner"
            )
    over_180 = sorted(key for key, stats in measured.items() if stats["max_s"] > 180)
    print(f"  Conditions with any measured attempt over 180 s: {len(over_180)} {over_180}")

    print()
    print("=" * 78)
    print("4. Failure modes from the 비고 notes")
    print("=" * 78)
    counts_by_model = {model: Counter() for model in model_order}
    counts_by_task = {task: Counter() for task in task_order}
    for trial in trials:
        for tag in trial["tags"]:
            counts_by_model[trial["model"]][tag] += 1
            counts_by_task[trial["task"]][tag] += 1
    ordered_modes = sorted(
        policy_failure_modes,
        key=lambda mode: -sum(counts_by_model[m][mode] for m in model_order),
    )
    ordered_modes = [mode for mode in ordered_modes if sum(counts_by_model[m][mode] for m in model_order) > 0]
    print(f"  {'mode':<26}" + "".join(f"{model:>8}" for model in model_order))
    for mode in ordered_modes:
        print(f"  {mode:<26}" + "".join(f"{counts_by_model[model][mode]:>8}" for model in model_order))
    print(
        f"  operator_event (rig / operator, not policy): "
        f"{ {model: counts_by_model[model]['operator_event'] for model in model_order} }"
    )
    notes_with_text = [t for t in trials if t["note"]]
    print(f"  {len(notes_with_text)} of {len(trials)} trials carry a note.")
    print(f"  Unclassified fragments ({len(unmatched_fragments)}):")
    for fragment in sorted(set(unmatched_fragments)):
        print(f"    - {fragment}")

    print()
    print("=" * 78)
    print("5. Outputs")
    print("=" * 78)
    written = [
        write_trials_csv(trials),
        write_conditions_csv(conditions),
        plot_heatmap(conditions),
        plot_scaling(conditions),
        plot_success_intervals(conditions),
        plot_components(conditions),
        plot_failure_modes(counts_by_model, ordered_modes),
    ]
    tables_path = output_dir / "tables.md"
    tables_path.write_text(
        "<!-- Generated by claude_test/eval_score_analysis.py -- do not edit. -->\n\n"
        "## Score and success by condition\n\n"
        + format_condition_table(conditions)
        + "\n\n## Rubric components\n\n"
        + format_component_table(conditions)
        + "\n\n## Failure modes\n\n"
        + format_failure_table(counts_by_model, ordered_modes)
        + "\n",
        encoding="utf-8",
    )
    written.append(tables_path)
    for path in written:
        print(f"  {path.relative_to(repo_root)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
